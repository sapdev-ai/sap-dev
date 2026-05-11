"""Build the canonical spec_template.xlsx with a self-describing (Meta) Layout sheet.

Bilingual: --lang EN (default) writes spec_template.xlsx,
           --lang JA writes spec_template_JA.xlsx.

Translation strings live in spec_translations.py — adding a third language
(e.g. ZH) means adding "ZH" entries there and running this script with
--lang ZH. No code changes needed in this file.

The template:
  - 17 content sheets + (Meta) Layout. Cover, Interface Contract,
    Selection Screen (image), Selection Definition, Validation Rules,
    Processing Flow, Mapping (File In), Mapping (File Out), Supplement
    (text), Domains, Data Elements, Tables, Error Messages, Text Elements,
    Golden Tests, Dependencies, README.
  - Tables sheet uses normalised two-block layout (Metadata + Fields,
    joined by the Table FK column).
  - (Meta) Layout sheet maps every section to its output file via stable
    English keys, with anchor named ranges for robustness against row
    insertion. Localised columns (sheet_name, source_column_header,
    anchor_keyword) come from spec_translations.py.
  - Four `format` values supported: kv, tsv, image, text.

The (Meta) Layout sheet is left VISIBLE in this build so a reviewer can
inspect it. In production, /sap-docs-layout sets it to hidden.

Usage:
  python tools/build_spec_template.py
  python tools/build_spec_template.py --lang JA
"""
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

from spec_translations import get_lang_strings, SUPPORTED_LANGS

# ---------------------------------------------------------------------------
# Parse --lang flag
# ---------------------------------------------------------------------------
LANG = "EN"
if "--lang" in sys.argv:
    i = sys.argv.index("--lang")
    if i + 1 < len(sys.argv):
        LANG = sys.argv[i + 1].upper()
if LANG not in SUPPORTED_LANGS:
    raise SystemExit(f"Unsupported --lang {LANG}. Use one of: {', '.join(SUPPORTED_LANGS)}")

L = get_lang_strings(LANG)
S, T, B, SB, H, K = L["S"], L["T"], L["B"], L["SB"], L["H"], L["K"]
COVER_LBL = L["COVER_LBL"]
README_LINES = L["README"]


# ---------------------------------------------------------------------------
# Output location
# ---------------------------------------------------------------------------
TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "plugins" / "sap-dev-core" / "shared" / "templates"
OUT = (TEMPLATES_DIR / "spec_template.xlsx" if LANG == "EN" else
       TEMPLATES_DIR / f"spec_template_{LANG}.xlsx")


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
FONT_TITLE   = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_HEADER  = Font(name="Calibri", size=11, bold=True)
FONT_NORMAL  = Font(name="Calibri", size=11)
FONT_BANNER  = Font(name="Calibri", size=11, bold=True, color="9C0006")

FILL_TITLE    = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION  = PatternFill("solid", fgColor="2E75B6")
FILL_HEADER   = PatternFill("solid", fgColor="DDEBF7")
FILL_BANNER   = PatternFill("solid", fgColor="FFF2CC")
FILL_META_KEY = PatternFill("solid", fgColor="E7E6E6")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title_row(ws, row, last_col, text):
    ws.cell(row=row, column=1, value=text).font = FONT_TITLE
    ws.cell(row=row, column=1).fill = FILL_TITLE
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.row_dimensions[row].height = 22


def style_section_row(ws, row, last_col, text):
    ws.cell(row=row, column=1, value=text).font = FONT_SECTION
    ws.cell(row=row, column=1).fill = FILL_SECTION
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.row_dimensions[row].height = 18


def style_header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 22


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_named_range(wb, name, sheet_title, ref):
    """Add a workbook-scoped named range pointing at sheet_title!ref."""
    safe_sheet = sheet_title.replace("'", "''")
    value = f"'{safe_sheet}'!${ref}"
    wb.defined_names[name] = DefinedName(name=name, attr_text=value)


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()
wb.remove(wb.active)


# 1. Cover --------------------------------------------------------------------
ws = wb.create_sheet(S["cover"])
style_title_row(ws, 1, 4, T["cover"])
ws.cell(row=2, column=1, value=B["cover"]).font = FONT_BANNER
ws.cell(row=2, column=1).fill = FILL_BANNER
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

style_header_row(ws, 3, H["cover"])
for i, label in enumerate(COVER_LBL, start=4):
    ws.cell(row=i, column=1, value=label).font = FONT_NORMAL
    ws.cell(row=i, column=1).border = BORDER_ALL
    ws.cell(row=i, column=2).border = BORDER_ALL
set_widths(ws, [28, 50, 4, 4])
ws.freeze_panes = "A4"
add_named_range(wb, "Cover.Anchor", S["cover"], "A3")


# 2. Interface Contract -------------------------------------------------------
ws = wb.create_sheet(S["interface"])
style_title_row(ws, 1, 12, T["interface"])

# Inputs A:D, Outputs E:H, Exceptions I:L
ws.cell(row=3, column=1, value=SB["iface_inputs"]).font = FONT_SECTION
ws.cell(row=3, column=1).fill = FILL_SECTION
ws.cell(row=3, column=1).alignment = ALIGN_LEFT
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

ws.cell(row=3, column=5, value=SB["iface_outputs"]).font = FONT_SECTION
ws.cell(row=3, column=5).fill = FILL_SECTION
ws.cell(row=3, column=5).alignment = ALIGN_LEFT
ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=8)

ws.cell(row=3, column=9, value=SB["iface_exceptions"]).font = FONT_SECTION
ws.cell(row=3, column=9).fill = FILL_SECTION
ws.cell(row=3, column=9).alignment = ALIGN_LEFT
ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=12)
ws.row_dimensions[3].height = 18

style_header_row(ws, 4, H["iface_inputs"] + H["iface_outputs"] + H["iface_exceptions"])
set_widths(ws, [16, 12, 8, 32, 16, 12, 8, 32, 12, 8, 24, 24])
ws.freeze_panes = "A5"
add_named_range(wb, "Interface.Inputs.Anchor",     S["interface"], "A4")
add_named_range(wb, "Interface.Outputs.Anchor",    S["interface"], "E4")
add_named_range(wb, "Interface.Exceptions.Anchor", S["interface"], "I4")


# 3. Selection Screen (image format) -----------------------------------------
# No header row, no TSV columns. Customer pastes a single image of the WHOLE
# selection screen below row 2. /sap-docs-extract saves the first embedded
# image as _selection_screen_layout.png.
ws = wb.create_sheet(S["selscr"])
style_title_row(ws, 1, 8, T["selscr"])
ws.cell(row=2, column=1, value=B["selscr"]).font = FONT_BANNER
ws.cell(row=2, column=1).fill = FILL_BANNER
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
# Wide single column so a typical screenshot fits without horizontal scrolling.
set_widths(ws, [120])
add_named_range(wb, "SelScr.Anchor", S["selscr"], "A1")


# 4. Selection Definition ----------------------------------------------------
ws = wb.create_sheet(S["seldef"])
style_title_row(ws, 1, len(H["seldef"]), T["seldef"])
ws.cell(row=2, column=1, value=B["seldef"]).font = FONT_BANNER
ws.cell(row=2, column=1).fill = FILL_BANNER
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(H["seldef"]))
style_header_row(ws, 3, H["seldef"])
set_widths(ws, [6, 18, 22, 18, 18, 12, 14, 14, 12, 16, 12, 30, 16])
ws.freeze_panes = "A4"
add_named_range(wb, "SelDef.Anchor", S["seldef"], "A3")


# 5. Validation Rules ---------------------------------------------------------
ws = wb.create_sheet(S["validation"])
style_title_row(ws, 1, 4, T["validation"])
style_header_row(ws, 3, H["validation"])
set_widths(ws, [6, 20, 64, 18])
ws.freeze_panes = "A4"
add_named_range(wb, "Validation.Anchor", S["validation"], "A3")


# 6. Processing Flow ----------------------------------------------------------
ws = wb.create_sheet(S["process"])
style_title_row(ws, 1, 3, T["process"])
style_header_row(ws, 3, H["process"])
set_widths(ws, [10, 50, 40])
ws.freeze_panes = "A4"
add_named_range(wb, "Process.Anchor", S["process"], "A3")


# 7. Mapping (File In) -------------------------------------------------------
ws = wb.create_sheet(S["filemap_in"])
style_title_row(ws, 1, len(H["filemap"]), T["filemap_in"])
ws.cell(row=2, column=1, value=B["filemap_in"]).font = FONT_BANNER
ws.cell(row=2, column=1).fill = FILL_BANNER
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(H["filemap"]))
style_header_row(ws, 3, H["filemap"])
set_widths(ws, [6, 22, 14, 8, 16, 16, 30, 14, 14])
ws.freeze_panes = "A4"
add_named_range(wb, "FileMap.In.Anchor", S["filemap_in"], "A3")


# 8. Mapping (File Out) ------------------------------------------------------
# Forward-looking skeleton for outbound. Same column structure as filemap_in.
ws = wb.create_sheet(S["filemap_out"])
style_title_row(ws, 1, len(H["filemap"]), T["filemap_out"])
ws.cell(row=2, column=1, value=B["filemap_out"]).font = FONT_BANNER
ws.cell(row=2, column=1).fill = FILL_BANNER
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(H["filemap"]))
style_header_row(ws, 3, H["filemap"])
set_widths(ws, [6, 22, 14, 8, 16, 16, 30, 14, 14])
ws.freeze_panes = "A4"
add_named_range(wb, "FileMap.Out.Anchor", S["filemap_out"], "A3")


# 9. Supplement (text format) ------------------------------------------------
# Free-form notes. No header row, no TSV columns. /sap-docs-extract dumps
# every non-empty cell from the sheet (row by row, tab-joined) as plain
# text into _supplement.txt.
ws = wb.create_sheet(S["supplement"])
style_title_row(ws, 1, 4, T["supplement"])
ws.cell(row=2, column=1, value=B["supplement"]).font = FONT_BANNER
ws.cell(row=2, column=1).fill = FILL_BANNER
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
set_widths(ws, [80, 30, 30, 30])
ws.freeze_panes = "A3"
add_named_range(wb, "Supplement.Anchor", S["supplement"], "A1")


# 10a. Domains ----------------------------------------------------------------
ws = wb.create_sheet(S["domains"])
style_title_row(ws, 1, 9, T["domains"])
ws.cell(row=2, column=1, value=B["domains"]).font = FONT_BANNER
ws.cell(row=2, column=1).fill = FILL_BANNER
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
style_header_row(ws, 3, H["domains"])
set_widths(ws, [20, 28, 14, 10, 10, 8, 12, 14, 18])
ws.freeze_panes = "A4"
add_named_range(wb, "DDIC.Domains.Anchor", S["domains"], "A3")


# 10b. Data Elements ----------------------------------------------------------
ws = wb.create_sheet(S["dataelements"])
style_title_row(ws, 1, 7, T["dataelements"])
ws.cell(row=2, column=1, value=B["dataelements"]).font = FONT_BANNER
ws.cell(row=2, column=1).fill = FILL_BANNER
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
style_header_row(ws, 3, H["dataelements"])
set_widths(ws, [22, 28, 18, 16, 18, 22, 22])
ws.freeze_panes = "A4"
add_named_range(wb, "DDIC.DataElements.Anchor", S["dataelements"], "A3")


# 10c. Tables -----------------------------------------------------------------
ws = wb.create_sheet(S["tables"])
style_title_row(ws, 1, 8, T["tables"])
style_section_row(ws, 3, 8, SB["tables_metadata"])
style_header_row(ws, 4, H["tables_metadata"])
style_section_row(ws, 17, 8, SB["tables_fields"])
style_header_row(ws, 18, H["tables_fields"])
set_widths(ws, [16, 6, 18, 8, 8, 18, 16, 14])
ws.freeze_panes = "A2"
add_named_range(wb, "DDIC.Tables.Metadata.Anchor", S["tables"], "A4")
add_named_range(wb, "DDIC.Tables.Fields.Anchor",   S["tables"], "A18")


# 11. Error Messages ----------------------------------------------------------
ws = wb.create_sheet(S["errmsgs"])
style_title_row(ws, 1, 4, T["errmsgs"])
style_header_row(ws, 3, H["errmsgs"])
set_widths(ws, [14, 10, 12, 60])
ws.freeze_panes = "A4"
add_named_range(wb, "ErrMsgs.Anchor", S["errmsgs"], "A3")


# 12. Text Elements -----------------------------------------------------------
ws = wb.create_sheet(S["textels"])
style_title_row(ws, 1, 2, T["textels"])
style_header_row(ws, 3, H["textels"])
set_widths(ws, [12, 60])
ws.freeze_panes = "A4"
add_named_range(wb, "TextEls.Anchor", S["textels"], "A3")


# 13. Golden Tests ------------------------------------------------------------
ws = wb.create_sheet(S["golden"])
style_title_row(ws, 1, 5, T["golden"])
style_header_row(ws, 3, H["golden"])
set_widths(ws, [10, 30, 40, 40, 30])
ws.freeze_panes = "A4"
add_named_range(wb, "Golden.Anchor", S["golden"], "A3")


# 14. Dependencies ------------------------------------------------------------
ws = wb.create_sheet(S["deps"])
style_title_row(ws, 1, 4, T["deps"])
style_header_row(ws, 3, H["deps"])
set_widths(ws, [12, 30, 40, 40])
ws.freeze_panes = "A4"
add_named_range(wb, "Deps.Anchor", S["deps"], "A3")


# 15. README ------------------------------------------------------------------
ws = wb.create_sheet(S["readme"])
style_title_row(ws, 1, 1, T["readme"])
for i, line in enumerate(README_LINES, start=2):
    ws.cell(row=i, column=1, value=line).font = FONT_NORMAL
set_widths(ws, [110])


# ===========================================================================
# (Meta) Layout — the contract sheet
# ===========================================================================
ws = wb.create_sheet(S["meta"])

# Header banner
style_title_row(ws, 1, 11, T["meta"])

ws.cell(row=2, column=1, value="schema_version").font = FONT_HEADER
ws.cell(row=2, column=1).fill = FILL_META_KEY
ws.cell(row=2, column=2, value=1).font = FONT_NORMAL

ws.cell(row=3, column=1, value="bootstrapped_from").font = FONT_HEADER
ws.cell(row=3, column=1).fill = FILL_META_KEY
ws.cell(row=3, column=2, value="<seeded by /sap-docs-layout bootstrap>").font = FONT_NORMAL

# Language row — declared in the workbook itself so /sap-docs-layout validate
# can warn on language/sheet-name mismatch.
ws.cell(row=4, column=1, value="language").font = FONT_HEADER
ws.cell(row=4, column=1).fill = FILL_META_KEY
ws.cell(row=4, column=2, value=LANG).font = FONT_NORMAL


# ----- ## SECTIONS table ----------------------------------------------------
ws.cell(row=6, column=1, value="## SECTIONS").font = FONT_BANNER
ws.cell(row=6, column=1).fill = FILL_BANNER

SECTIONS_HEADERS = [
    "key", "sheet_name", "section_label", "output_file", "format",
    "anchor_named_range", "anchor_keyword", "header_row_offset",
    "data_starts_offset", "required", "notes",
]
style_header_row(ws, 7, SECTIONS_HEADERS)

# section_label is kept stable English in V1 (consumed by the skill UI). If
# customer feedback shows JA users want translated labels in `inspect`
# output, add a T_LABELS dict to spec_translations.py and switch here.
SECTIONS_ROWS = [
    # Order matches the workbook tab order so /sap-docs-layout inspect
    # produces a logical walkthrough.
    # key                       sheet_name              section_label              output_file                      format    anchor_named_range                  anchor_keyword               hdr  data req     notes
    ("cover",                   S["cover"],             "Program Summary",         "_PGM_summary.txt",              "kv",     "Cover.Anchor",                     K["cover"],                  0,   1,   "yes",  "One field per row, two columns"),
    ("interface_inputs",        S["interface"],         "Interface Inputs",        "_interface.txt#inputs",         "tsv",    "Interface.Inputs.Anchor",          K["iface_inputs"],           0,   1,   "yes",  ""),
    ("interface_outputs",       S["interface"],         "Interface Outputs",       "_interface.txt#outputs",        "tsv",    "Interface.Outputs.Anchor",         K["iface_outputs"],          0,   1,   "yes",  ""),
    ("interface_exceptions",    S["interface"],         "Interface Exceptions",    "_interface.txt#exceptions",     "tsv",    "Interface.Exceptions.Anchor",      K["iface_exceptions"],       0,   1,   "no",   ""),
    ("selscr",                  S["selscr"],            "Selection Screen Layout", "_selection_screen_layout.png",  "image",  "SelScr.Anchor",                    K["selscr"],                 0,   0,   "no",   "Whole-screen image; no COLUMNS rows; consumed by /sap-gen-abap as multimodal input"),
    ("seldef",                  S["seldef"],            "Selection Definition",    "_selection_definition.txt",     "tsv",    "SelDef.Anchor",                    K["seldef"],                 0,   1,   "no",   "Source of truth for PARAMETERS / SELECT-OPTIONS"),
    ("validation",              S["validation"],        "Validation Rules",        "_process.txt#validation",       "tsv",    "Validation.Anchor",                K["validation"],             0,   1,   "no",   "Folded into _process.txt"),
    ("process",                 S["process"],           "Processing Flow",         "_process.txt#process",          "tsv",    "Process.Anchor",                   K["process"],                0,   1,   "yes",  "Folded into _process.txt"),
    ("filemap_in",              S["filemap_in"],        "Mapping (File In)",       "_file_mapping_in.txt",          "tsv",    "FileMap.In.Anchor",                K["filemap_in"],             0,   1,   "no",   "File field -> SAP table.field"),
    ("filemap_out",             S["filemap_out"],       "Mapping (File Out)",      "_file_mapping_out.txt",         "tsv",    "FileMap.Out.Anchor",               K["filemap_out"],            0,   1,   "no",   "SAP table.field -> file field; future use"),
    ("supplement",              S["supplement"],        "Supplement",              "_supplement.txt",               "text",   "Supplement.Anchor",                K["supplement"],             0,   0,   "no",   "Free-form notes; whole sheet dumped verbatim; no COLUMNS rows"),
    ("ddic_domains",            S["domains"],           "DDIC Domains",            "_domains.txt",                  "tsv",    "DDIC.Domains.Anchor",              K["domains"],                0,   1,   "yes",  ""),
    ("ddic_dataelements",       S["dataelements"],      "DDIC Data Elements",      "_dataElements.txt",             "tsv",    "DDIC.DataElements.Anchor",         K["dataelements"],           0,   1,   "yes",  ""),
    ("ddic_tables_metadata",    S["tables"],            "DDIC Tables (metadata)",  "_tables.txt",                   "tsv",    "DDIC.Tables.Metadata.Anchor",      K["tables_metadata"],        0,   1,   "yes",  "Joined with ddic_tables_fields by Table column"),
    ("ddic_tables_fields",      S["tables"],            "DDIC Tables (fields)",    "_tables.txt",                   "tsv",    "DDIC.Tables.Fields.Anchor",        K["tables_fields"],          0,   1,   "yes",  "Joined with ddic_tables_metadata by Table column"),
    ("errmsgs",                 S["errmsgs"],           "Error Messages",          "_errorMsgs.txt",                "tsv",    "ErrMsgs.Anchor",                   K["errmsgs"],                0,   1,   "no",   ""),
    ("textels",                 S["textels"],           "Text Elements",           "_textElements.txt",             "tsv",    "TextEls.Anchor",                   K["textels"],                0,   1,   "no",   ""),
    ("golden",                  S["golden"],            "Golden Tests",            "_golden.txt",                   "tsv",    "Golden.Anchor",                    K["golden"],                 0,   1,   "no",   ""),
    ("deps",                    S["deps"],              "Dependencies",            "_deps.txt",                     "tsv",    "Deps.Anchor",                      K["deps"],                   0,   1,   "no",   ""),
]
for i, row_data in enumerate(SECTIONS_ROWS, start=8):
    for j, val in enumerate(row_data, start=1):
        c = ws.cell(row=i, column=j, value=val)
        c.font = FONT_NORMAL
        c.border = BORDER_ALL
        c.alignment = ALIGN_LEFT

next_row = 8 + len(SECTIONS_ROWS) + 2  # blank row gap

# ----- ## COLUMNS table -----------------------------------------------------
ws.cell(row=next_row, column=1, value="## COLUMNS").font = FONT_BANNER
ws.cell(row=next_row, column=1).fill = FILL_BANNER
header_row = next_row + 1

COLUMNS_HEADERS = [
    "section_key", "output_position", "output_column",
    "source_column_header", "source_column_letter", "transform",
    "default_if_blank", "required", "notes",
]
style_header_row(ws, header_row, COLUMNS_HEADERS + [""] * (len(SECTIONS_HEADERS) - len(COLUMNS_HEADERS)))
for j in range(len(COLUMNS_HEADERS) + 1, len(SECTIONS_HEADERS) + 1):
    ws.cell(row=header_row, column=j).fill = PatternFill()
    ws.cell(row=header_row, column=j).border = Border()


def col(section_key, position, output_column, source_section_key, source_index,
        letter, transform, default_blank, required, notes):
    """Build one COLUMN_ROWS row using H[source_section_key][source_index] for
    the localized source_column_header. Source section key may differ from
    section_key (e.g. cover header lookup uses H['cover']).
    """
    return (section_key, position, output_column,
            H[source_section_key][source_index], letter, transform,
            default_blank, required, notes)


COLUMN_ROWS = [
    # ----- cover -----
    # source_column_header is a meta-comment for kv format (parser doesn't scan).
    ("cover", 1, "FIELD", "(col A labels)", "A", "trim", "", "yes", "Field name column"),
    ("cover", 2, "VALUE", "(col B values)", "B", "trim", "", "yes", "Field value column"),

    # ----- interface_inputs -----
    col("interface_inputs", 1, "NAME",        "iface_inputs", 0, "A", "trim",  "", "yes", ""),
    col("interface_inputs", 2, "TYPE",        "iface_inputs", 1, "B", "upper", "", "yes", ""),
    col("interface_inputs", 3, "LENGTH",      "iface_inputs", 2, "C", "int",   "", "no",  ""),
    col("interface_inputs", 4, "DESCRIPTION", "iface_inputs", 3, "D", "trim",  "", "no",  ""),

    # ----- interface_outputs -----
    col("interface_outputs", 1, "NAME",        "iface_outputs", 0, "E", "trim",  "", "yes", ""),
    col("interface_outputs", 2, "TYPE",        "iface_outputs", 1, "F", "upper", "", "yes", ""),
    col("interface_outputs", 3, "LENGTH",      "iface_outputs", 2, "G", "int",   "", "no",  ""),
    col("interface_outputs", 4, "DESCRIPTION", "iface_outputs", 3, "H", "trim",  "", "no",  ""),

    # ----- interface_exceptions -----
    col("interface_exceptions", 1, "MSG_CLASS",   "iface_exceptions", 0, "I", "trim",  "", "yes", ""),
    col("interface_exceptions", 2, "MSG_NO",      "iface_exceptions", 1, "J", "trim",  "", "yes", ""),
    col("interface_exceptions", 3, "WHEN_RAISED", "iface_exceptions", 2, "K", "trim",  "", "no",  ""),
    col("interface_exceptions", 4, "ACTION",      "iface_exceptions", 3, "L", "trim",  "", "no",  ""),

    # ----- validation -----
    col("validation", 1, "NO",          "validation", 0, "A", "int",  "", "yes", ""),
    col("validation", 2, "FIELD",       "validation", 1, "B", "trim", "", "no",  ""),
    col("validation", 3, "RULE",        "validation", 2, "C", "trim", "", "yes", ""),
    col("validation", 4, "ERR_MSG_REF", "validation", 3, "D", "trim", "", "no",  ""),

    # ----- process -----
    col("process", 1, "STEP",   "process", 0, "A", "trim", "", "yes", ""),
    col("process", 2, "ACTION", "process", 1, "B", "trim", "", "yes", ""),
    col("process", 3, "NOTES",  "process", 2, "C", "trim", "", "no",  ""),

    # ----- ddic_domains -----
    col("ddic_domains", 1, "DOMAIN_NAME",   "domains", 0, "A", "upper",  "",  "yes", ""),
    col("ddic_domains", 2, "SHORT_DESC",    "domains", 1, "B", "trim",   "",  "yes", ""),
    col("ddic_domains", 3, "DATATYPE",      "domains", 2, "C", "upper",  "",  "yes", ""),
    col("ddic_domains", 4, "LENGTH",        "domains", 3, "D", "int",    "",  "yes", ""),
    col("ddic_domains", 5, "DECIMALS",      "domains", 4, "E", "int",    "0", "no",  ""),
    col("ddic_domains", 6, "SIGN",          "domains", 5, "F", "bool_X", "",  "no",  ""),
    col("ddic_domains", 7, "LOWERCASE",     "domains", 6, "G", "bool_X", "",  "no",  ""),
    col("ddic_domains", 8, "OUTPUT_LENGTH", "domains", 7, "H", "int",    "",  "no",  ""),
    col("ddic_domains", 9, "CONV_ROUTINE",  "domains", 8, "I", "trim",   "",  "no",  ""),

    # ----- ddic_dataelements -----
    col("ddic_dataelements", 1, "DTEL_NAME",     "dataelements", 0, "A", "upper", "", "yes", ""),
    col("ddic_dataelements", 2, "SHORT_DESC",    "dataelements", 1, "B", "trim",  "", "yes", ""),
    col("ddic_dataelements", 3, "DOMNAME",       "dataelements", 2, "C", "upper", "", "yes", ""),
    col("ddic_dataelements", 4, "LABEL_SHORT",   "dataelements", 3, "D", "trim",  "", "no",  ""),
    col("ddic_dataelements", 5, "LABEL_MEDIUM",  "dataelements", 4, "E", "trim",  "", "no",  ""),
    col("ddic_dataelements", 6, "LABEL_LONG",    "dataelements", 5, "F", "trim",  "", "no",  ""),
    col("ddic_dataelements", 7, "LABEL_HEADING", "dataelements", 6, "G", "trim",  "", "no",  ""),

    # ----- ddic_tables_metadata -----
    col("ddic_tables_metadata", 1, "TABLE_NAME",     "tables_metadata", 0, "A", "upper", "", "yes", ""),
    col("ddic_tables_metadata", 2, "SHORT_DESC",     "tables_metadata", 1, "B", "trim",  "", "yes", ""),
    col("ddic_tables_metadata", 3, "DELIVERY_CLASS", "tables_metadata", 2, "C", "upper", "", "yes", ""),
    col("ddic_tables_metadata", 4, "DATA_CLASS",     "tables_metadata", 3, "D", "upper", "", "no",  ""),
    col("ddic_tables_metadata", 5, "SIZE_CATEGORY",  "tables_metadata", 4, "E", "upper", "", "no",  ""),

    # ----- ddic_tables_fields -----
    col("ddic_tables_fields", 1, "TABLE",       "tables_fields", 0, "A", "upper",  "", "yes", "FK to ddic_tables_metadata.TABLE_NAME"),
    col("ddic_tables_fields", 2, "NO",          "tables_fields", 1, "B", "int",    "", "yes", ""),
    col("ddic_tables_fields", 3, "FIELDNAME",   "tables_fields", 2, "C", "upper",  "", "yes", ""),
    col("ddic_tables_fields", 4, "KEY",         "tables_fields", 3, "D", "bool_X", "", "no",  ""),
    col("ddic_tables_fields", 5, "INITIAL",     "tables_fields", 4, "E", "bool_X", "", "no",  ""),
    col("ddic_tables_fields", 6, "DATAELEMENT", "tables_fields", 5, "F", "upper",  "", "yes", ""),
    col("ddic_tables_fields", 7, "REFTABLE",    "tables_fields", 6, "G", "upper",  "", "no",  ""),
    col("ddic_tables_fields", 8, "REFFIELD",    "tables_fields", 7, "H", "upper",  "", "no",  ""),

    # ----- errmsgs -----
    col("errmsgs", 1, "MSG_CLASS", "errmsgs", 0, "A", "upper", "", "yes", ""),
    col("errmsgs", 2, "MSG_NO",    "errmsgs", 1, "B", "trim",  "", "yes", ""),
    col("errmsgs", 3, "MSG_TYPE",  "errmsgs", 2, "C", "upper", "", "yes", "Validates against E/W/I/S/A"),
    col("errmsgs", 4, "MSG_TEXT",  "errmsgs", 3, "D", "trim",  "", "yes", ""),

    # ----- textels -----
    col("textels", 1, "TEXT_ID",    "textels", 0, "A", "trim", "", "yes", ""),
    col("textels", 2, "TEXT_VALUE", "textels", 1, "B", "trim", "", "yes", ""),

    # ----- selscr -----
    # NO COLUMN ROWS — image format. /sap-docs-extract saves the first embedded
    # image from the sheet as _selection_screen_layout.png.

    # ----- supplement -----
    # NO COLUMN ROWS — text format. /sap-docs-extract dumps every non-empty
    # cell from the sheet (row by row, tab-joined) as plain text into
    # _supplement.txt.

    # ----- seldef -----
    col("seldef",  1, "NO",             "seldef", 0,  "A", "int",   "", "yes", ""),
    col("seldef",  2, "LABEL",          "seldef", 1,  "B", "trim",  "", "yes", ""),
    col("seldef",  3, "NAME_JA",        "seldef", 2,  "C", "trim",  "", "yes", ""),
    col("seldef",  4, "NAME_EN",        "seldef", 3,  "D", "upper", "", "no",  "ABAP parameter identifier"),
    col("seldef",  5, "DTEL_NAME",      "seldef", 4,  "E", "upper", "", "yes", ""),
    col("seldef",  6, "DATATYPE",       "seldef", 5,  "F", "upper", "", "no",  "Only when DTEL is a primitive type"),
    col("seldef",  7, "LENGTH",         "seldef", 6,  "G", "int",   "", "yes", ""),
    col("seldef",  8, "DECIMALS",       "seldef", 7,  "H", "int",   "0", "no",  ""),
    col("seldef",  9, "IO_TYPE",        "seldef", 8,  "I", "trim",  "", "yes", "INPUT / OUTPUT / BOTH"),
    col("seldef", 10, "DISPLAY_FORMAT", "seldef", 9,  "J", "trim",  "", "no",  ""),
    col("seldef", 11, "MANDATORY",      "seldef", 10, "K", "trim",  "", "yes", "Normalised by /sap-docs-convert: REQUIRED / OPTIONAL / CONDITIONAL"),
    col("seldef", 12, "DESCRIPTION",    "seldef", 11, "L", "trim",  "", "no",  ""),
    col("seldef", 13, "DEFAULT_VALUE",  "seldef", 12, "M", "trim",  "", "no",  ""),

    # ----- filemap_in -----
    col("filemap_in", 1, "NO",                "filemap", 0, "A", "int",   "", "yes", ""),
    col("filemap_in", 2, "FILE_FIELD",        "filemap", 1, "B", "trim",  "", "yes", ""),
    col("filemap_in", 3, "DATATYPE",          "filemap", 2, "C", "upper", "", "yes", ""),
    col("filemap_in", 4, "LENGTH",            "filemap", 3, "D", "int",   "", "yes", ""),
    col("filemap_in", 5, "MANDATORY_CREATE",  "filemap", 4, "E", "trim",  "", "yes", "Normalised: REQUIRED / OPTIONAL / CONDITIONAL / FIXED"),
    col("filemap_in", 6, "MANDATORY_UPDATE",  "filemap", 5, "F", "trim",  "", "yes", "Normalised: REQUIRED / OPTIONAL / CONDITIONAL / FIXED"),
    col("filemap_in", 7, "NOTES",             "filemap", 6, "G", "trim",  "", "no",  ""),
    col("filemap_in", 8, "SAP_TABLE",         "filemap", 7, "H", "upper", "", "no",  ""),
    col("filemap_in", 9, "SAP_FIELD",         "filemap", 8, "I", "upper", "", "no",  ""),

    # ----- filemap_out (same column shape as filemap_in) -----
    col("filemap_out", 1, "NO",               "filemap", 0, "A", "int",   "", "yes", ""),
    col("filemap_out", 2, "FILE_FIELD",       "filemap", 1, "B", "trim",  "", "yes", ""),
    col("filemap_out", 3, "DATATYPE",         "filemap", 2, "C", "upper", "", "yes", ""),
    col("filemap_out", 4, "LENGTH",           "filemap", 3, "D", "int",   "", "yes", ""),
    col("filemap_out", 5, "MANDATORY_CREATE", "filemap", 4, "E", "trim",  "", "no",  "Reserved for future outbound semantics"),
    col("filemap_out", 6, "MANDATORY_UPDATE", "filemap", 5, "F", "trim",  "", "no",  "Reserved for future outbound semantics"),
    col("filemap_out", 7, "NOTES",            "filemap", 6, "G", "trim",  "", "no",  ""),
    col("filemap_out", 8, "SAP_TABLE",        "filemap", 7, "H", "upper", "", "no",  ""),
    col("filemap_out", 9, "SAP_FIELD",        "filemap", 8, "I", "upper", "", "no",  ""),

    # ----- golden -----
    col("golden", 1, "TEST_ID",  "golden", 0, "A", "trim", "", "yes", ""),
    col("golden", 2, "SCENARIO", "golden", 1, "B", "trim", "", "yes", ""),
    col("golden", 3, "INPUTS",   "golden", 2, "C", "trim", "", "yes", ""),
    col("golden", 4, "EXPECTED", "golden", 3, "D", "trim", "", "yes", ""),
    col("golden", 5, "NOTES",    "golden", 4, "E", "trim", "", "no",  ""),

    # ----- deps -----
    col("deps", 1, "TYPE",    "deps", 0, "A", "upper", "", "yes", "FM / BAPI / INCLUDE / CLASS"),
    col("deps", 2, "NAME",    "deps", 1, "B", "upper", "", "yes", ""),
    col("deps", 3, "PURPOSE", "deps", 2, "C", "trim",  "", "no",  ""),
    col("deps", 4, "NOTES",   "deps", 3, "D", "trim",  "", "no",  ""),
]
for i, row_data in enumerate(COLUMN_ROWS, start=header_row + 1):
    for j, val in enumerate(row_data, start=1):
        c = ws.cell(row=i, column=j, value=val)
        c.font = FONT_NORMAL
        c.border = BORDER_ALL
        c.alignment = ALIGN_LEFT

set_widths(ws, [22, 22, 22, 26, 16, 12, 16, 10, 28, 0, 0])
ws.freeze_panes = "A8"
# Leave visible in this preview build; production /sap-docs-layout sets hidden.
# ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
wb.save(OUT)
print(f"WROTE [{LANG}]: {OUT}")
print(f"  Sheets: {len(wb.sheetnames)}")
for name in wb.sheetnames:
    marker = " (visible - preview only)" if name == S["meta"] else ""
    print(f"    - {name}{marker}")
print(f"  Named ranges: {len(wb.defined_names)}")
print(f"  (Meta) Layout sections: {len(SECTIONS_ROWS)}")
print(f"  (Meta) Layout column rows: {len(COLUMN_ROWS)}")
