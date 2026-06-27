"""Simulate /sap-docs-extract against a workbook with a (Meta) Layout sheet.

Walks the SECTIONS rows, dispatches by format (kv / tsv / image / text),
and writes the same _*.txt / _*.png files the production skill would.

This is a TEST harness: it lets us run the agent's pipeline end-to-end
without an LLM-driven skill execution, surfacing any gap in the schema or
in /sap-docs-extract's instructions.
"""
import sys
import io
from datetime import datetime
from pathlib import Path
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage

if len(sys.argv) < 2:
    raise SystemExit("usage: test_extract_simulation.py <spec.xlsx>")
SPEC = Path(sys.argv[1]).resolve()
if not SPEC.exists():
    raise SystemExit(f"not found: {SPEC}")

# Settings defaults (Step 0.1)
WORK_DIR = Path(r"C:\sap_dev_work")
SOURCE_CODE_URL = WORK_DIR / "source_code"

# Derive doc_name and work_folder (Step 1c)
doc_name = SPEC.stem
if len(doc_name) > 50:
    doc_name = doc_name[:50]
ts = datetime.now().strftime("%Y%m%d%H%M%S")
work_folder = SOURCE_CODE_URL / "work" / f"{doc_name}_{ts}"
work_folder.mkdir(parents=True, exist_ok=True)
print(f"INFO: work_folder = {work_folder}")

# Step 1c — dump xlsx to _raw.txt
raw_path = work_folder / f"{doc_name}_raw.txt"
wb = load_workbook(SPEC, data_only=True)
lines = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    lines.append(f"========== Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ==========")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
        if any(c is not None for c in row):
            cells = ["" if c is None else str(c) for c in row]
            lines.append(f"R{row_idx}\t" + "\t".join(cells))
    lines.append("")
raw_path.write_text("\n".join(lines), encoding="utf-8")
print(f"INFO: _raw.txt written ({raw_path.stat().st_size:,} bytes)")

# Step 2 — read (Meta) Layout
if "(Meta) Layout" not in wb.sheetnames:
    raise SystemExit("ERROR: workbook has no (Meta) Layout sheet")
meta = wb["(Meta) Layout"]

# Find ## SECTIONS marker, header row, then walk rows
def find_marker(ws, marker):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == marker:
            return r
    return None

sections_marker = find_marker(meta, "## SECTIONS")
columns_marker = find_marker(meta, "## COLUMNS")
if not sections_marker or not columns_marker:
    raise SystemExit("ERROR: Meta sheet missing ## SECTIONS or ## COLUMNS marker")

# Header row is right after the marker
sections_header_row = sections_marker + 1
columns_header_row = columns_marker + 1

def parse_table(ws, header_row, end_row):
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if not h:
            break
        headers.append(str(h))
    rows = []
    for r in range(header_row + 1, end_row):
        row = {}
        for c, h in enumerate(headers, start=1):
            v = ws.cell(row=r, column=c).value
            row[h] = "" if v is None else str(v)
        if any(v.strip() for v in row.values()):
            rows.append(row)
    return rows

sections = parse_table(meta, sections_header_row, columns_marker)
columns = parse_table(meta, columns_header_row, meta.max_row + 1)
print(f"INFO: parsed {len(sections)} sections, {len(columns)} column rows")

# Group columns by section_key
cols_by_section = {}
for c in columns:
    cols_by_section.setdefault(c["section_key"], []).append(c)

# Resolve named-range anchors
named_ranges = {}
for n in wb.defined_names:
    dn = wb.defined_names[n]
    named_ranges[n] = dn.attr_text

def resolve_anchor(section):
    """Returns (sheet_name, row_index_1based) for the anchor cell, or (sheet_name, None)."""
    nr = section.get("anchor_named_range", "").strip()
    if nr and nr in named_ranges:
        ref = named_ranges[nr]
        # Format: 'SheetName'!$A3 or 'SheetName'!$A$3
        m = re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)", ref)
        if m:
            sheet, col, row = m.group(1), m.group(2), int(m.group(3))
            # Honour the named range ONLY when it resolves to a sheet that
            # actually exists in THIS workbook. A range copied between
            # workbooks is rewritten by Excel into an external reference like
            # '[4]SourceTemplateSheet'!$A18 — the sheet does not exist here,
            # so it is NOT resolvable. Fall through to the keyword scan
            # instead of reading the section's real sheet at the wrong row.
            if sheet in wb.sheetnames:
                return sheet, row, column_index_from_string(col)
    # Fallback: scan ALL columns for the keyword (top-to-bottom, then
    # left-to-right within each row). The anchor keyword may live in any
    # column — e.g. ddic_tables_fields uses FIELDNAME, which sits in column C
    # (after the Table join column in A and No in B), so a column-A-only scan
    # never finds it and the fields half of _tables.txt comes out empty.
    sheet = section.get("sheet_name", "")
    keyword = section.get("anchor_keyword", "")
    if sheet and keyword and sheet in wb.sheetnames:
        ws_local = wb[sheet]
        for r in range(1, ws_local.max_row + 1):
            for c in range(1, ws_local.max_column + 1):
                v = ws_local.cell(row=r, column=c).value
                if v and str(v).strip() == keyword:
                    return sheet, r, c
    return sheet, None, 1

# Per-format dispatch
output_summary = []

def transform_value(v, transform):
    s = str(v) if v is not None else ""
    if transform == "trim":
        return s.strip()
    elif transform == "upper":
        return s.strip().upper()
    elif transform == "lower":
        return s.strip().lower()
    elif transform == "int":
        s = s.strip()
        try:
            return str(int(float(s))) if s else ""
        except ValueError:
            return s
    elif transform.startswith("bool_"):
        true_marker = transform.split("_", 1)[1]
        return "X" if s.strip() == true_marker else ""
    return s.strip()

def apply_kv(section, ws, anchor_row):
    """For Cover sheet: anchor is FIELD/VALUE header row. Data rows below."""
    out_lines = []
    if anchor_row is None:
        return None
    # Header at anchor_row, data from anchor_row + 1
    for r in range(anchor_row + 1, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if key:
            out_lines.append(f"{str(key).strip()}\t{str(val or '').strip()}")
    return "\n".join(out_lines)

def apply_tsv(section, ws, anchor_row, anchor_col, cols):
    if anchor_row is None or not cols:
        return None
    # Output header
    sorted_cols = sorted(cols, key=lambda c: int(c.get("output_position") or 0))
    out_header = [c["output_column"] for c in sorted_cols]
    # Data starts at anchor_row + data_starts_offset (default 1)
    data_offset = int(section.get("data_starts_offset") or 1)
    data_start = anchor_row + data_offset
    out_rows = [out_header]
    # Determine bounds — read from data_start until a fully-empty row
    for r in range(data_start, ws.max_row + 1):
        row_vals = []
        any_non_empty = False
        for c in sorted_cols:
            letter = c.get("source_column_letter", "").strip()
            if not letter:
                row_vals.append("")
                continue
            col_idx = column_index_from_string(letter)
            v = ws.cell(row=r, column=col_idx).value
            if v is not None and str(v).strip():
                any_non_empty = True
            transformed = transform_value(v, c.get("transform", "trim"))
            row_vals.append(transformed)
        if not any_non_empty:
            break
        out_rows.append(row_vals)
    return "\n".join("\t".join(row) for row in out_rows)

def apply_text(section, ws):
    """Free-form: dump every non-empty cell row by row, tab-joined, starting row 3."""
    lines = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        cells = ["" if c is None else str(c) for c in row]
        while cells and cells[-1] == "":
            cells.pop()
        lines.append("\t".join(cells))
    while lines and lines[-1].strip() == "":
        lines.pop()
    return "\n".join(lines)

def apply_image(section, ws, output_path):
    """Save first embedded image."""
    if not ws._images:
        return False
    img = ws._images[0]
    try:
        pil = PILImage.open(io.BytesIO(img._data()))
        pil.save(output_path, format="PNG")
        return True
    except Exception as e:
        print(f"WARN: image save failed for {section['key']}: {e}")
        return False

# Track sections that share output_file (e.g., interface_inputs/outputs/exceptions all → _interface.txt)
file_buffers = {}

# Per-output-file tracker: last non-blank value seen in column 1 across all
# sections that write to this file. Lets a downstream TSV section inherit
# its FK / join column from an upstream section when the spec author left
# the FK column blank on every row. Without this, a single-table spec like
# `MaterialUpload_JA.xlsx` produces `_tables.txt` whose fields sub-section
# has blank TABLE column on every row — even though the metadata sub-section
# above has the table name (`ZMMFIXEDVALS30`). Tracker key = output_file
# path; value = last non-blank column-1 string. See SKILL.md Step 3.
fk_carry = {}


def forward_fill_first_column(tsv_text, carry_initial=""):
    """Walk a TSV string top-down, forward-filling blank column 1 cells.

    First line is the header row — never filled, but its column-1 value is
    NOT used as a carry source (headers describe the schema, not data).
    Returns (new_tsv_text, last_non_blank_col1_value).
    """
    if not tsv_text:
        return tsv_text, carry_initial
    lines = tsv_text.split("\n")
    if len(lines) < 2:
        return tsv_text, carry_initial
    out = [lines[0]]
    carry = carry_initial
    for raw in lines[1:]:
        if not raw.strip():
            out.append(raw)
            continue
        cells = raw.split("\t")
        if cells and not cells[0].strip() and carry:
            cells[0] = carry
            out.append("\t".join(cells))
        else:
            if cells and cells[0].strip():
                carry = cells[0].strip()
            out.append(raw)
    return "\n".join(out), carry


for section in sections:
    fmt = section.get("format", "").strip()
    sheet_name = section.get("sheet_name", "").strip()
    output_file = section.get("output_file", "").strip()
    key = section.get("key", "").strip()
    if not sheet_name or not output_file:
        continue
    if sheet_name not in wb.sheetnames:
        print(f"WARN: section {key}: sheet {sheet_name!r} not found in workbook")
        continue
    ws = wb[sheet_name]

    sheet, anchor_row, anchor_col = resolve_anchor(section)

    # Strip URL fragment from output_file (e.g. _interface.txt#inputs → _interface.txt)
    file_part, _, frag = output_file.partition("#")
    out_path = work_folder / f"{doc_name}{file_part}"

    content = None
    if fmt == "kv":
        content = apply_kv(section, ws, anchor_row)
    elif fmt == "tsv":
        cols = cols_by_section.get(key, [])
        content = apply_tsv(section, ws, anchor_row, anchor_col, cols)
        # Forward-fill the join key (column 1) — both within this section
        # (carry from any non-blank value already in the section) and
        # across sections that share this output_file (e.g. tables_metadata
        # → tables_fields). The header row is never filled.
        if content:
            content, fk_carry[out_path] = forward_fill_first_column(
                content, fk_carry.get(out_path, "")
            )
    elif fmt == "image":
        ok = apply_image(section, ws, out_path)
        output_summary.append(f"  {key:24s} → {out_path.name} {'(image)' if ok else '(NO IMAGE)'}")
        continue
    elif fmt == "text":
        content = apply_text(section, ws)
    else:
        print(f"WARN: section {key}: unknown format {fmt!r}")
        continue

    if content is None:
        print(f"WARN: section {key}: no content produced")
        continue

    # Append to file (use frag as a separator if multi-section file)
    label = f"== {frag.upper()} ==" if frag else None
    file_buffers.setdefault(out_path, [])
    if label:
        file_buffers[out_path].append(label)
    file_buffers[out_path].append(content)
    output_summary.append(f"  {key:24s} → {out_path.name}#{frag}" if frag else f"  {key:24s} → {out_path.name}")

# Flush buffers
for path, parts in file_buffers.items():
    path.write_text("\n\n".join(parts) + "\n", encoding="utf-8")

print()
print("=== Output files ===")
for line in output_summary:
    print(line)
print()
print(f"Work folder: {work_folder}")

# List actual files written
print()
print("=== Files on disk ===")
for f in sorted(work_folder.iterdir()):
    print(f"  {f.name}  ({f.stat().st_size:,} bytes)")
