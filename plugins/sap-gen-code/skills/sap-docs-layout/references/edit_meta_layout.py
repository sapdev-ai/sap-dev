#!/usr/bin/env python3
"""edit_meta_layout.py -- openpyxl helper for /sap-docs-layout.

All workbook (.xlsx) logic for the sap-docs-layout skill lives here: reading,
validating and structurally editing the "(Meta) Layout" sheet that
/sap-docs-extract reads as the per-workbook layout source of truth.
Schema doc: <SKILL_DIR>/templates/spec_layout_schema.md.

CLI:
  python edit_meta_layout.py inspect      <workbook>
  python edit_meta_layout.py bootstrap    <workbook> [--from <source.xlsx>] [--force]
  python edit_meta_layout.py add-column   <workbook> --section <key> --name <COL>
                                          --after <existing|START>
                                          [--source-header <text>] [--required]
                                          [--transform trim|upper|int|bool_X]
  python edit_meta_layout.py rename-sheet <workbook> --section <key> --to <new-name>
  python edit_meta_layout.py validate     <workbook> [--dry-parse]

Exit codes:
  0  success (validate: Phase A finished with no FAIL findings)
  1  validate finished and found FAIL-level findings
  2  hard error -- an "ERROR: <CLASS>: <message>" line names the error class
     (WORKBOOK_NOT_FOUND, META_NOT_FOUND, META_ALREADY_EXISTS, SCHEMA_UNSUPPORTED,
      SECTION_UNKNOWN, SHEET_NOT_FOUND, ANCHOR_NOT_FOUND, INVALID_ARGUMENT,
      XLSX_WRITE_FAILED, PYTHON_DEPENDENCY, UNEXPECTED)

Guarantees:
  * Content-sheet DATA cells are never touched. Only sheet names, column
    structure (add-column), header cells, named ranges and the (Meta) Layout
    sheet itself are in scope.
  * The helper never backs up the workbook -- the SKILL.md flow (Step 2)
    writes a timestamped .bak copy before any write operation.
  * Write operations fail loud (non-zero exit + ERROR line), never silently.
"""

import argparse
import datetime
import os
import re
import sys
from copy import copy

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter, quote_sheetname
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("ERROR: PYTHON_DEPENDENCY: openpyxl is required by /sap-docs-layout. "
          "Install it with: pip install openpyxl")
    sys.exit(2)

META_SHEET = "(Meta) Layout"
SUPPORTED_SCHEMA_VERSIONS = (1,)
SECTIONS_BANNER = "## SECTIONS"
COLUMNS_BANNER = "## COLUMNS"
SECTIONS_HEADERS = [
    "key", "sheet_name", "section_label", "output_file", "format",
    "anchor_named_range", "anchor_keyword", "header_row_offset",
    "data_starts_offset", "required", "notes",
]
COLUMNS_HEADERS = [
    "section_key", "output_position", "output_column", "source_column_header",
    "source_column_letter", "transform", "default_if_blank", "required", "notes",
]
KNOWN_FORMATS = ("kv", "tsv", "image", "text")
ADD_COLUMN_TRANSFORMS = ("trim", "upper", "int", "bool_X")
# Sheets exempt from the orphan check (SKILL.md validate contract).
ORPHAN_EXEMPT_PREFIX = "(Auto) "


class MetaError(Exception):
    """Hard error with a stable error class for the SKILL.md Log-End step."""

    def __init__(self, error_class, message):
        super().__init__(message)
        self.error_class = error_class


def _s(value):
    """Cell value -> trimmed string ('' for None)."""
    if value is None:
        return ""
    return str(value).strip()


def _i(value, default=0):
    """Cell value -> int (tolerates '3', 3, '' -> default)."""
    s = _s(value)
    if s == "":
        return default
    try:
        return int(float(s))
    except ValueError:
        return default


# ---------------------------------------------------------------------------
# Meta-sheet parsing
# ---------------------------------------------------------------------------

def read_meta(wb, workbook_label=""):
    """Parse the (Meta) Layout sheet into a dict.

    Returns {ws, header, header_rows, sections, sections_by_key, columns,
             sections_banner_row, sections_header_row, sections_col_idx,
             columns_banner_row, columns_header_row, columns_col_idx}.
    Raises MetaError(META_NOT_FOUND / SCHEMA_UNSUPPORTED / INVALID_ARGUMENT).
    """
    if META_SHEET not in wb.sheetnames:
        raise MetaError("META_NOT_FOUND",
                        "workbook %s has no '%s' sheet. Run: /sap-docs-layout bootstrap <workbook>"
                        % (workbook_label or "<target>", META_SHEET))
    ws = wb[META_SHEET]

    sections_banner_row = None
    columns_banner_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = _s(row[0].value)
        if v == SECTIONS_BANNER and sections_banner_row is None:
            sections_banner_row = row[0].row
        elif v == COLUMNS_BANNER and columns_banner_row is None:
            columns_banner_row = row[0].row
    if sections_banner_row is None or columns_banner_row is None:
        raise MetaError("META_NOT_FOUND",
                        "'%s' sheet is present but missing its '%s' / '%s' banner rows -- "
                        "not a spec_layout_schema meta sheet" % (META_SHEET, SECTIONS_BANNER, COLUMNS_BANNER))

    # Header key/value rows live between the title row and the SECTIONS banner.
    header = {}
    header_rows = {}
    for r in range(2, sections_banner_row):
        k = _s(ws.cell(row=r, column=1).value)
        if k and not k.startswith("##"):
            header[k] = _s(ws.cell(row=r, column=2).value)
            header_rows[k] = r

    schema_version = _i(header.get("schema_version"), default=-1)
    if schema_version not in SUPPORTED_SCHEMA_VERSIONS:
        raise MetaError("SCHEMA_UNSUPPORTED",
                        "meta sheet schema_version '%s' is not supported by this helper (supported: %s). "
                        "Refusing to operate." % (header.get("schema_version"),
                                                  ", ".join(str(v) for v in SUPPORTED_SCHEMA_VERSIONS)))

    def read_table(header_row, expected_headers, stop_row):
        col_idx = {}
        c = 1
        while True:
            name = _s(ws.cell(row=header_row, column=c).value)
            if name == "":
                break
            col_idx[name] = c
            c += 1
        missing = [h for h in expected_headers if h not in col_idx]
        if missing:
            raise MetaError("INVALID_ARGUMENT",
                            "meta table at row %d is missing expected header column(s): %s"
                            % (header_row, ", ".join(missing)))
        rows = []
        r = header_row + 1
        while r <= ws.max_row and (stop_row is None or r < stop_row):
            first = _s(ws.cell(row=r, column=col_idx[expected_headers[0]]).value)
            if first == "" or first.startswith("##"):
                break
            rec = {"_row": r}
            for name, c in col_idx.items():
                rec[name] = _s(ws.cell(row=r, column=c).value)
            rows.append(rec)
            r += 1
        return rows, col_idx

    sections_header_row = sections_banner_row + 1
    sections, sections_col_idx = read_table(sections_header_row, SECTIONS_HEADERS, columns_banner_row)
    columns_header_row = columns_banner_row + 1
    columns, columns_col_idx = read_table(columns_header_row, COLUMNS_HEADERS, None)

    sections_by_key = {}
    for sec in sections:
        sections_by_key[sec["key"]] = sec

    return {
        "ws": ws,
        "header": header,
        "header_rows": header_rows,
        "sections": sections,
        "sections_by_key": sections_by_key,
        "columns": columns,
        "sections_banner_row": sections_banner_row,
        "sections_header_row": sections_header_row,
        "sections_col_idx": sections_col_idx,
        "columns_banner_row": columns_banner_row,
        "columns_header_row": columns_header_row,
        "columns_col_idx": columns_col_idx,
    }


def section_columns(meta, key):
    cols = [c for c in meta["columns"] if c["section_key"] == key]
    cols.sort(key=lambda c: _i(c["output_position"]))
    return cols


# ---------------------------------------------------------------------------
# Anchor resolution (mirrors /sap-docs-extract semantics)
# ---------------------------------------------------------------------------

_COORD_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)([0-9]+)$")


def _parse_coord(coord):
    """'$A18' / 'B4' -> (row, col) or None."""
    m = _COORD_RE.match(coord.strip())
    if not m:
        return None
    return int(m.group(4)), column_index_from_string(m.group(2).upper())


def resolve_named_range(wb, name, expect_sheet=None):
    """Resolve a single-cell named range -> (sheet, row, col) or None.

    Follows the /sap-docs-extract rule: usable ONLY when it resolves to a sheet
    that exists in THIS workbook (external '[N]Sheet' refs fall through), and,
    when expect_sheet is given, to that sheet.
    """
    if not name:
        return None
    try:
        dn = wb.defined_names.get(name)
    except Exception:
        dn = None
    if dn is None:
        return None
    try:
        dests = list(dn.destinations)
    except Exception:
        return None
    for sheet, coord in dests:
        if sheet not in wb.sheetnames:
            continue
        if expect_sheet is not None and sheet != expect_sheet:
            continue
        first = coord.split(":")[0]
        rc = _parse_coord(first)
        if rc:
            return sheet, rc[0], rc[1]
    return None


def find_anchor_by_keyword(ws, keyword):
    """First cell whose trimmed value equals keyword, scanning ALL columns
    top-to-bottom then left-to-right within each row (the /sap-docs-extract
    scan order -- NOT just column A). Returns (row, col) or None."""
    kw = _s(keyword)
    if kw == "":
        return None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and _s(cell.value) == kw:
                return cell.row, cell.column
    return None


def resolve_section_anchor(wb, ws, sec):
    """Named range first, keyword scan second. Returns (row, col, how) or None."""
    nr = resolve_named_range(wb, sec["anchor_named_range"], expect_sheet=ws.title)
    if nr:
        return nr[1], nr[2], "named_range"
    kw = find_anchor_by_keyword(ws, sec["anchor_keyword"])
    if kw:
        return kw[0], kw[1], "keyword"
    return None


# ---------------------------------------------------------------------------
# inspect
# ---------------------------------------------------------------------------

def op_inspect(args):
    wb = _load(args.workbook)
    try:
        meta = read_meta(wb, args.workbook)
    except MetaError as e:
        if e.error_class == "META_NOT_FOUND":
            print("WARN: This workbook has no (Meta) Layout sheet. Run:")
            print("  /sap-docs-layout bootstrap %s" % args.workbook)
            print("to create one from the built-in defaults.")
            return 0
        raise

    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(args.workbook))
    print("Workbook : %s" % os.path.abspath(args.workbook))
    print("Modified : %s" % mtime.strftime("%Y-%m-%d %H:%M:%S"))
    print("Sheets   : %d (meta sheet state: %s)" % (len(wb.sheetnames), meta["ws"].sheet_state))
    for k in ("schema_version", "language", "bootstrapped_from"):
        if k in meta["header"]:
            print("Meta     : %s = %s" % (k, meta["header"][k]))
    print("")
    print("## Sections")
    print("")
    print("| key | sheet_name | output_file | format | required |")
    print("|---|---|---|---|---|")
    for sec in meta["sections"]:
        print("| %s | %s | %s | %s | %s |" % (
            sec["key"], sec["sheet_name"], sec["output_file"], sec["format"], sec["required"]))
    print("")
    print("## Columns")
    total_cols = 0
    for sec in meta["sections"]:
        cols = section_columns(meta, sec["key"])
        total_cols += len(cols)
        if not cols:
            continue
        print("")
        print("### %s" % sec["key"])
        print("")
        print("| output_position | output_column | source_column_header | source_column_letter | transform | required |")
        print("|---|---|---|---|---|---|")
        for c in cols:
            print("| %s | %s | %s | %s | %s | %s |" % (
                c["output_position"], c["output_column"], c["source_column_header"],
                c["source_column_letter"], c["transform"], c["required"]))
    required_count = sum(1 for s in meta["sections"] if s["required"].lower() == "yes")
    print("")
    print("%d sections, %d columns, %d required sections"
          % (len(meta["sections"]), total_cols, required_count))
    return 0


# ---------------------------------------------------------------------------
# bootstrap
# ---------------------------------------------------------------------------

def default_bootstrap_source():
    """Canonical shared template, resolved plugin-relative from this file:
    references/ -> skill -> skills -> plugin -> plugins root -> sap-dev-core."""
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.normpath(os.path.join(
        here, "..", "..", "..", "..", "sap-dev-core", "shared", "templates", "spec_template.xlsx"))


def op_bootstrap(args):
    source = args.source or default_bootstrap_source()
    if not os.path.isfile(source):
        raise MetaError("WORKBOOK_NOT_FOUND",
                        "bootstrap source workbook not found: %s (pass --from <path>)" % source)
    swb = _load(source)
    if META_SHEET not in swb.sheetnames:
        raise MetaError("META_NOT_FOUND", "bootstrap source %s has no '%s' sheet" % (source, META_SHEET))
    sws = swb[META_SHEET]

    twb = _load(args.workbook)
    if META_SHEET in twb.sheetnames:
        if not args.force:
            raise MetaError("META_ALREADY_EXISTS",
                            "(Meta) Layout already exists. Use --force to overwrite or run\n"
                            "  /sap-docs-layout inspect %s\n"
                            "to review the current layout first." % args.workbook)
        del twb[META_SHEET]

    tws = twb.create_sheet(META_SHEET)
    for row in sws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            t = tws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                t.font = copy(cell.font)
                t.fill = copy(cell.fill)
                t.border = copy(cell.border)
                t.alignment = copy(cell.alignment)
                t.number_format = cell.number_format
    for key, dim in sws.column_dimensions.items():
        if dim.width is not None:
            tws.column_dimensions[key].width = dim.width
    try:
        for mr in list(getattr(sws, "merged_cells").ranges):
            tws.merge_cells(str(mr))
    except Exception:
        pass
    tws.freeze_panes = sws.freeze_panes
    # Sheet-scoped defined names (none in the canonical template; best-effort).
    try:
        for name, dn in getattr(sws, "defined_names", {}).items():
            tws.defined_names[name] = DefinedName(name=name, attr_text=dn.attr_text)
    except Exception as e:
        print("WARN: could not copy sheet-scoped named ranges: %s" % e)

    # Stamp provenance + language sanity.
    meta = read_meta(twb, args.workbook)
    r = meta["header_rows"].get("bootstrapped_from")
    if r:
        tws.cell(row=r, column=2, value="%s @ %s" % (
            os.path.basename(source), datetime.date.today().isoformat()))
    lang = meta["header"].get("language", "")
    localized = [n for n in twb.sheetnames
                 if n != META_SHEET and any(ord(ch) > 127 for ch in n)]
    if localized and lang.upper() == "EN":
        print("WARN: target workbook has localized sheet name(s) (e.g. '%s') but the meta "
              "'language' row says EN -- verify/adjust the language row in (Meta) Layout."
              % localized[0])

    tws.sheet_state = "hidden"
    _save(twb, args.workbook)
    print("OK: (Meta) Layout bootstrapped into %s" % args.workbook)
    print("    from %s (%d section rows, %d column rows). Sheet is hidden."
          % (source, len(meta["sections"]), len(meta["columns"])))
    print("    Run /sap-docs-layout validate %s to confirm." % args.workbook)
    return 0


# ---------------------------------------------------------------------------
# add-column
# ---------------------------------------------------------------------------

def _shift_defined_names_for_col_insert(wb, sheet_name, insert_idx):
    """After ws.insert_cols(insert_idx) on sheet_name, shift every defined-name
    destination on that sheet at column >= insert_idx one column right
    (openpyxl does not re-point references on structural edits)."""
    updated = 0
    for name in list(wb.defined_names.keys()):
        dn = wb.defined_names[name]
        try:
            dests = list(dn.destinations)
        except Exception:
            continue
        if not dests or all(sheet != sheet_name for sheet, _ in dests):
            continue
        parts = []
        changed = False
        for sheet, coord in dests:
            new_coord = coord
            if sheet == sheet_name:
                pieces = []
                for piece in coord.split(":"):
                    m = _COORD_RE.match(piece.strip())
                    if not m:
                        pieces = None
                        break
                    row_n = int(m.group(4))
                    col_n = column_index_from_string(m.group(2).upper())
                    if col_n >= insert_idx:
                        col_n += 1
                        changed = True
                    pieces.append("%s%s%s%d" % (m.group(1), get_column_letter(col_n), m.group(3), row_n))
                if pieces is None:
                    print("WARN: named range '%s' has an unparseable reference (%s); not adjusted." % (name, coord))
                    pieces = [coord]
                new_coord = ":".join(pieces)
            parts.append("%s!%s" % (quote_sheetname(sheet), new_coord))
        if changed:
            wb.defined_names[name] = DefinedName(name=name, attr_text=",".join(parts))
            updated += 1
    return updated


def op_add_column(args):
    wb = _load(args.workbook)
    meta = read_meta(wb, args.workbook)
    key = args.section
    if key not in meta["sections_by_key"]:
        raise MetaError("SECTION_UNKNOWN", "unknown section '%s'. Known keys: %s"
                        % (key, ", ".join(s["key"] for s in meta["sections"])))
    sec = meta["sections_by_key"][key]
    fmt = sec["format"]
    if fmt not in ("kv", "tsv"):
        raise MetaError("INVALID_ARGUMENT",
                        "section '%s' has format '%s'; only kv/tsv sections have columns" % (key, fmt))
    if args.transform not in ADD_COLUMN_TRANSFORMS:
        raise MetaError("INVALID_ARGUMENT", "unknown --transform '%s'. Allowed: %s"
                        % (args.transform, ", ".join(ADD_COLUMN_TRANSFORMS)))
    new_name = _s(args.name)
    if new_name == "" or any(ch in new_name for ch in "\t|"):
        raise MetaError("INVALID_ARGUMENT", "--name must be a non-empty column name without tabs/pipes")

    cols = section_columns(meta, key)
    if not cols:
        raise MetaError("INVALID_ARGUMENT",
                        "section '%s' has no existing COLUMNS rows to anchor against" % key)
    if any(c["output_column"] == new_name for c in cols):
        raise MetaError("INVALID_ARGUMENT",
                        "section '%s' already has an output column named '%s'" % (key, new_name))

    if args.after.upper() == "START":
        new_position = 1
        ws_insert_idx = min(column_index_from_string(c["source_column_letter"]) for c in cols
                            if c["source_column_letter"])
        meta_insert_row = cols[0]["_row"]
    else:
        anchor_col = None
        for c in cols:
            if c["output_column"] == args.after:
                anchor_col = c
                break
        if anchor_col is None:
            raise MetaError("INVALID_ARGUMENT",
                            "--after '%s' is not an output column of section '%s'. Use START or one of: %s"
                            % (args.after, key, ", ".join(c["output_column"] for c in cols)))
        new_position = _i(anchor_col["output_position"]) + 1
        if not anchor_col["source_column_letter"]:
            raise MetaError("INVALID_ARGUMENT",
                            "column '%s' has no source_column_letter; cannot compute the insert position"
                            % args.after)
        ws_insert_idx = column_index_from_string(anchor_col["source_column_letter"]) + 1
        meta_insert_row = anchor_col["_row"] + 1

    sheet_name = sec["sheet_name"]
    if sheet_name not in wb.sheetnames:
        raise MetaError("SHEET_NOT_FOUND", "section '%s' points at sheet '%s' which does not exist"
                        % (key, sheet_name))
    ws = wb[sheet_name]
    anchor = resolve_section_anchor(wb, ws, sec)
    if anchor is None:
        raise MetaError("ANCHOR_NOT_FOUND",
                        "cannot locate the header row of section '%s' (named range '%s' unresolvable "
                        "and keyword '%s' not found on sheet '%s')"
                        % (key, sec["anchor_named_range"], sec["anchor_keyword"], sheet_name))
    header_row = anchor[0] + _i(sec["header_row_offset"])

    # 1. Structural edit on the content sheet (header only -- no data cells).
    ws.insert_cols(ws_insert_idx)
    source_header = args.source_header if args.source_header is not None else new_name
    hcell = ws.cell(row=header_row, column=ws_insert_idx, value=source_header)
    neighbor_idx = ws_insert_idx - 1 if ws_insert_idx > 1 else ws_insert_idx + 1
    neighbor = ws.cell(row=header_row, column=neighbor_idx)
    if neighbor.has_style:
        hcell.font = copy(neighbor.font)
        hcell.fill = copy(neighbor.fill)
        hcell.border = copy(neighbor.border)
        hcell.alignment = copy(neighbor.alignment)

    # 2. Re-point named ranges on the same sheet (openpyxl does not).
    nr_updated = _shift_defined_names_for_col_insert(wb, sheet_name, ws_insert_idx)

    # 3. Meta updates.
    mws = meta["ws"]
    ci = meta["columns_col_idx"]
    mws.insert_rows(meta_insert_row)

    def actual_row(cached_row):
        return cached_row + 1 if cached_row >= meta_insert_row else cached_row

    # 3a. Shift source_column_letter of every column row on the SAME sheet
    #     (any section) sitting at/right of the inserted worksheet column.
    letters_shifted = 0
    same_sheet_keys = set(s["key"] for s in meta["sections"] if s["sheet_name"] == sheet_name)
    for c in meta["columns"]:
        if c["section_key"] not in same_sheet_keys or not c["source_column_letter"]:
            continue
        idx = column_index_from_string(c["source_column_letter"])
        if idx >= ws_insert_idx:
            mws.cell(row=actual_row(c["_row"]), column=ci["source_column_letter"],
                     value=get_column_letter(idx + 1))
            letters_shifted += 1

    # 3b. Renumber output_position inside the target section.
    for c in cols:
        if _i(c["output_position"]) >= new_position:
            mws.cell(row=actual_row(c["_row"]), column=ci["output_position"],
                     value=_i(c["output_position"]) + 1)

    # 3c. Write the new meta row (style copied from the row below = old neighbor).
    style_src_row = meta_insert_row + 1
    values = {
        "section_key": key,
        "output_position": new_position,
        "output_column": new_name,
        "source_column_header": source_header,
        "source_column_letter": get_column_letter(ws_insert_idx),
        "transform": args.transform,
        "default_if_blank": "",
        "required": "yes" if args.required else "no",
        "notes": "",
    }
    for name, cidx in ci.items():
        cell = mws.cell(row=meta_insert_row, column=cidx, value=values.get(name, ""))
        src = mws.cell(row=style_src_row, column=cidx)
        if src.has_style:
            cell.font = copy(src.font)
            cell.fill = copy(src.fill)
            cell.border = copy(src.border)
            cell.alignment = copy(src.alignment)

    _save(wb, args.workbook)
    print("OK: Added column %s to sheet \"%s\" at position %d (worksheet column %s)."
          % (new_name, sheet_name, new_position, get_column_letter(ws_insert_idx)))
    print("    Meta updated: section=%s, output_position=%d; %d same-sheet letter(s) shifted; "
          "%d named range(s) re-pointed." % (key, new_position, letters_shifted, nr_updated))
    print("    Run /sap-docs-layout validate %s to confirm." % args.workbook)
    return 0


# ---------------------------------------------------------------------------
# rename-sheet
# ---------------------------------------------------------------------------

def op_rename_sheet(args):
    wb = _load(args.workbook)
    meta = read_meta(wb, args.workbook)
    key = args.section
    if key not in meta["sections_by_key"]:
        raise MetaError("SECTION_UNKNOWN", "unknown section '%s'. Known keys: %s"
                        % (key, ", ".join(s["key"] for s in meta["sections"])))
    old = meta["sections_by_key"][key]["sheet_name"]
    if old not in wb.sheetnames:
        raise MetaError("SHEET_NOT_FOUND", "section '%s' points at sheet '%s' which does not exist"
                        % (key, old))
    new = args.to
    if not (1 <= len(new) <= 31):
        raise MetaError("INVALID_ARGUMENT", "new sheet name must be 1-31 characters (got %d)" % len(new))
    bad = [ch for ch in ":\\/?*[]" if ch in new]
    if bad:
        raise MetaError("INVALID_ARGUMENT", "new sheet name contains forbidden character(s): %s"
                        % " ".join(bad))
    if new == old:
        raise MetaError("INVALID_ARGUMENT", "sheet is already named '%s'" % old)
    for existing in wb.sheetnames:
        if existing != old and existing.lower() == new.lower():
            raise MetaError("INVALID_ARGUMENT",
                            "a sheet named '%s' already exists in this workbook" % existing)

    wb[old].title = new

    # Meta: sections can SHARE a sheet (Interface Contract, Tables) -- update
    # every section row mapped to the renamed sheet, not just the named one.
    mws = meta["ws"]
    si = meta["sections_col_idx"]
    sections_updated = 0
    for sec in meta["sections"]:
        if sec["sheet_name"] == old:
            mws.cell(row=sec["_row"], column=si["sheet_name"], value=new)
            sections_updated += 1

    # Named ranges: openpyxl does not rewrite refs on rename -- do it here.
    nr_updated = 0
    for name in list(wb.defined_names.keys()):
        dn = wb.defined_names[name]
        try:
            dests = list(dn.destinations)
        except Exception:
            continue
        if not dests or all(sheet != old for sheet, _ in dests):
            continue
        parts = []
        for sheet, coord in dests:
            target = new if sheet == old else sheet
            parts.append("%s!%s" % (quote_sheetname(target), coord))
        wb.defined_names[name] = DefinedName(name=name, attr_text=",".join(parts))
        nr_updated += 1

    _save(wb, args.workbook)
    print("OK: Renamed sheet \"%s\" -> \"%s\"." % (old, new))
    print("    Meta section row(s) updated: %d. Named range(s) updated: %d."
          % (sections_updated, nr_updated))
    return 0


# ---------------------------------------------------------------------------
# validate (Phase A -- structural reconciliation; Phase B dry-parse is
# orchestrated by the SKILL.md flow via /sap-docs-extract, not by this helper)
# ---------------------------------------------------------------------------

def op_validate(args):
    wb = _load(args.workbook)
    meta = read_meta(wb, args.workbook)
    findings = []  # (level, message)

    def warn(msg):
        findings.append(("WARN", msg))

    def fail(msg):
        findings.append(("FAIL", msg))

    columns_checked = 0
    for sec in meta["sections"]:
        key = sec["key"]
        fmt = sec["format"]
        sheet_name = sec["sheet_name"]
        cols = section_columns(meta, key)

        if fmt not in KNOWN_FORMATS:
            fail("section \"%s\": unknown format '%s'. Allowed: kv, tsv, image, text." % (key, fmt))
            continue
        if sheet_name not in wb.sheetnames:
            fail("section \"%s\": sheet '%s' not found in the workbook" % (key, sheet_name))
            continue
        ws = wb[sheet_name]

        if fmt in ("kv", "tsv"):
            min_cols = 2 if fmt == "kv" else 1
            if len(cols) < min_cols:
                fail("section \"%s\": format '%s' needs at least %d COLUMNS row(s), found %d"
                     % (key, fmt, min_cols, len(cols)))
            anchor = resolve_section_anchor(wb, ws, sec)
            anchor_row = None
            if anchor is None:
                warn("section \"%s\": anchor not found (named range '%s' unresolvable and keyword "
                     "'%s' not found on any column of sheet '%s')"
                     % (key, sec["anchor_named_range"], sec["anchor_keyword"], sheet_name))
            else:
                anchor_row = anchor[0]
            header_row = None
            if anchor_row is not None:
                header_row = anchor_row + _i(sec["header_row_offset"])

            for c in cols:
                columns_checked += 1
                letter_ok = False
                cidx = None
                if c["source_column_letter"]:
                    try:
                        cidx = column_index_from_string(c["source_column_letter"])
                        letter_ok = cidx <= ws.max_column
                    except ValueError:
                        letter_ok = False
                header_found = False
                if not letter_ok and header_row is not None and c["source_column_header"]:
                    for cell in ws[header_row]:
                        if cell.value is not None and _s(cell.value) == c["source_column_header"]:
                            header_found = True
                            break
                if not letter_ok and not header_found:
                    warn("section \"%s\": column \"%s\": source_column_letter '%s' out of range and "
                         "source_column_header '%s' not found on the anchor row"
                         % (key, c["output_column"], c["source_column_letter"], c["source_column_header"]))
                # Required columns must carry a real header cell. tsv only: kv
                # sections have no on-sheet header row (their source_column_header
                # values are meta-comments like "(col A labels)").
                if fmt == "tsv" and c["required"].lower() == "yes":
                    if header_row is None or cidx is None:
                        warn("section \"%s\": column \"%s\" is required but its header cell could not "
                             "be located (no anchor / no valid column letter)" % (key, c["output_column"]))
                    elif _s(ws.cell(row=header_row, column=cidx).value) == "":
                        fail("section \"%s\": column \"%s\" required but header cell %s%d is blank"
                             % (key, c["output_column"], c["source_column_letter"], header_row))

        elif fmt == "image":
            if len(cols) != 0:
                fail("section \"%s\": image-format sections must not have COLUMNS entries (found %d) -- "
                     "they describe a single embedded image, not tabular data" % (key, len(cols)))
            if len(getattr(ws, "_images", [])) == 0:
                warn("section \"%s\": sheet '%s' has no embedded image yet -- paste a whole-screen "
                     "image into the sheet before extraction" % (key, sheet_name))
            if sec["anchor_named_range"] and resolve_named_range(
                    wb, sec["anchor_named_range"], expect_sheet=sheet_name) is None:
                warn("section \"%s\": anchor_named_range '%s' does not resolve on sheet '%s'"
                     % (key, sec["anchor_named_range"], sheet_name))
            if sec["anchor_keyword"]:
                warn("section \"%s\": anchor_keyword '%s' is set but keyword scanning is meaningless "
                     "for image-format sections -- leave it empty" % (key, sec["anchor_keyword"]))

        elif fmt == "text":
            if len(cols) != 0:
                fail("section \"%s\": text-format sections are unstructured and must not have COLUMNS "
                     "entries (found %d)" % (key, len(cols)))
            if sec["anchor_named_range"] and resolve_named_range(
                    wb, sec["anchor_named_range"], expect_sheet=sheet_name) is None:
                warn("section \"%s\": anchor_named_range '%s' does not resolve on sheet '%s'"
                     % (key, sec["anchor_named_range"], sheet_name))
            if sec["anchor_keyword"]:
                warn("section \"%s\": anchor_keyword '%s' is set but not used for text-format "
                     "sections -- leave it empty" % (key, sec["anchor_keyword"]))
            # An empty sheet is acceptable (customer opted out) -- no warning.

    # Orphan sheets: not meta, not README, not "(Auto) *", not referenced.
    referenced = set(s["sheet_name"] for s in meta["sections"])
    for name in wb.sheetnames:
        if name == META_SHEET or name.upper() == "README" or name.startswith(ORPHAN_EXEMPT_PREFIX):
            continue
        if name not in referenced:
            warn("orphan sheet '%s' is not referenced by any meta section" % name)

    warn_count = sum(1 for lv, _ in findings if lv == "WARN")
    fail_count = sum(1 for lv, _ in findings if lv == "FAIL")
    print("=== Phase A: Structural reconciliation ===")
    print("OK   %d sections checked" % len(meta["sections"]))
    print("OK   %d columns checked" % columns_checked)
    print("WARN %d warnings" % warn_count)
    print("FAIL %d errors" % fail_count)
    if args.dry_parse:
        print("")
        print("INFO: Phase B (dry-parse) is orchestrated by the /sap-docs-layout skill "
              "(it runs /sap-docs-extract); this helper performs Phase A only.")
    print("")
    print("=== Punch list ===")
    if findings:
        for lv, msg in findings:
            print("- [%s] %s" % (lv, msg))
    else:
        print("- (none)")
    return 1 if fail_count > 0 else 0


# ---------------------------------------------------------------------------
# plumbing
# ---------------------------------------------------------------------------

def _load(path):
    if not os.path.isfile(path):
        raise MetaError("WORKBOOK_NOT_FOUND", "workbook not found: %s" % path)
    try:
        return openpyxl.load_workbook(path)
    except Exception as e:
        raise MetaError("INVALID_ARGUMENT", "cannot open %s as an xlsx workbook: %s" % (path, e))


def _save(wb, path):
    try:
        wb.save(path)
    except Exception as e:
        raise MetaError("XLSX_WRITE_FAILED", "could not write %s: %s (is it open in Excel?)" % (path, e))


def build_parser():
    p = argparse.ArgumentParser(
        prog="edit_meta_layout.py",
        description="openpyxl helper for /sap-docs-layout -- read/validate/edit the (Meta) Layout sheet.")
    sub = p.add_subparsers(dest="op", required=True)

    sp = sub.add_parser("inspect", help="print the current layout")
    sp.add_argument("workbook")
    sp.set_defaults(func=op_inspect)

    sp = sub.add_parser("bootstrap", help="copy a (Meta) Layout sheet from a reference workbook")
    sp.add_argument("workbook")
    sp.add_argument("--from", dest="source", default=None,
                    help="source workbook (default: the canonical shared spec_template.xlsx)")
    sp.add_argument("--force", action="store_true", help="overwrite an existing (Meta) Layout sheet")
    sp.set_defaults(func=op_bootstrap)

    sp = sub.add_parser("add-column", help="add a column to a kv/tsv section (xlsx + meta in one step)")
    sp.add_argument("workbook")
    sp.add_argument("--section", required=True)
    sp.add_argument("--name", required=True)
    sp.add_argument("--after", required=True, help="existing output column to insert after, or START")
    sp.add_argument("--source-header", dest="source_header", default=None)
    sp.add_argument("--required", action="store_true")
    sp.add_argument("--transform", default="trim", help="one of: %s" % ", ".join(ADD_COLUMN_TRANSFORMS))
    sp.set_defaults(func=op_add_column)

    sp = sub.add_parser("rename-sheet", help="rename a sheet and update meta + named ranges")
    sp.add_argument("workbook")
    sp.add_argument("--section", required=True)
    sp.add_argument("--to", required=True)
    sp.set_defaults(func=op_rename_sheet)

    sp = sub.add_parser("validate", help="reconcile the meta sheet against the workbook (Phase A)")
    sp.add_argument("workbook")
    sp.add_argument("--dry-parse", dest="dry_parse", action="store_true",
                    help="acknowledged flag; the dry-parse itself is run by the skill")
    sp.set_defaults(func=op_validate)
    return p


def main(argv=None):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    args = build_parser().parse_args(argv)
    try:
        return args.func(args)
    except MetaError as e:
        print("ERROR: %s: %s" % (e.error_class, e))
        return 2
    except Exception as e:
        print("ERROR: UNEXPECTED: %s: %s" % (type(e).__name__, e))
        return 2


if __name__ == "__main__":
    sys.exit(main())
